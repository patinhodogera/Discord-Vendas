import discord
from discord import app_commands
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os

class VendaLogCog(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.Cog.listener()
    async def on_ready(self):
        print(f'{__name__} foi carregado')

    async def canal_autocomplete(self, interaction: discord.Interaction, current: str):
        guild = interaction.guild
        return [
            app_commands.Choice(name=channel.name, value=str(channel.id))
            for channel in guild.text_channels
            if current.lower() in channel.name.lower()
        ]

    @app_commands.command(name="gerar_planilha_vendas", description="Gera uma planilha Excel com os dados das vendas do canal selecionado")
    @app_commands.autocomplete(canal=canal_autocomplete)
    async def gerar_planilha_vendas(self, interaction: discord.Interaction, canal: str):
        """Busca todas as vendas anteriores do canal de logs e gera uma planilha Excel"""
        
        channel = interaction.guild.get_channel(int(canal))
        if not channel:
            await interaction.response.send_message("Canal não encontrado.", ephemeral=True)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Vendas Realizadas"

        header_fill = PatternFill(start_color="FFB3C1E1", end_color="FFB3C1E1", fill_type="solid") 
        data_fill_even = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        lucro_fill = PatternFill(start_color="FF5CB85C", end_color="FF5CB85C", fill_type="solid")
        font_header = Font(bold=True, size=12, color="FF000000")
        font_lucro = Font(bold=True, size=12, color="FFFFFFFF")
        thin_border = Border(left=Side(style='thin', color="FFAAAAAA"), 
                             right=Side(style='thin', color="FFAAAAAA"), 
                             top=Side(style='thin', color="FFAAAAAA"), 
                             bottom=Side(style='thin', color="FFAAAAAA"))
        alignment_center = Alignment(horizontal="center", vertical="center")

        headers = ["Comprador", "Data", "Produto", "Valor Pago"]
        ws.append(headers)

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.fill = header_fill
                cell.font = font_header 
                cell.border = thin_border 
                cell.alignment = alignment_center 

        total_lucro = 0.0

        row_num = 2
        async for message in channel.history(limit=None):
            for embed in message.embeds:
                if embed.title == "Venda Concluída":
                    comprador = embed.fields[0].value
                    data = embed.fields[1].value
                    produto = embed.fields[2].value
                    valor_pago = float(embed.fields[3].value.replace('R$', '').replace(',', '.').strip())
                    total_lucro += valor_pago
                    ws.append([comprador, data, produto, f"R$ {valor_pago:.2f}"])
                    for col in ws.iter_cols(min_row=row_num, max_row=row_num, min_col=1, max_col=4):
                        for cell in col:
                            cell.border = thin_border
                            cell.alignment = alignment_center
                            if row_num % 2 == 0:
                                cell.fill = data_fill_even
                    row_num += 1
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 20

        ws.append([])
        ws.append(["", "", "Lucro Total:", f"R$ {total_lucro:.2f}"])

        lucro_cell = ws.cell(row=ws.max_row, column=4)
        lucro_cell.fill = lucro_fill
        lucro_cell.font = font_lucro 
        lucro_cell.border = thin_border
        lucro_cell.alignment = alignment_center

        arquivo_excel = "vendas_realizadas.xlsx"
        wb.save(arquivo_excel)

        await interaction.response.send_message("Aqui está a planilha com as vendas realizadas:", file=discord.File(arquivo_excel))
        os.remove(arquivo_excel)

async def setup(bot):
    await bot.add_cog(VendaLogCog(bot))
